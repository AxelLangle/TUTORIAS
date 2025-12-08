"""
Script para cargar datos de prueba desde el archivo Excel
"""

import sqlite3
import openpyxl
from datetime import datetime

DATABASE = 'asesorias.db'
EXCEL_FILE = '/home/ubuntu/upload/Datosdepruebabasededatostutorias.xlsx'

def load_data():
    """Carga los datos de estudiantes desde el archivo Excel"""
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        estudiantes_insertados = 0
        estudiantes_duplicados = 0
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\nProcesando hoja: {sheet_name}")
            
            # Extraer grupo de la hoja (ej. "1725 IS" -> grupo_id)
            grupo_id = sheet_name.replace(" ", "")
            
            # Decodificar grupo para obtener cuatrimestre y carrera
            if len(grupo_id) >= 5:
                grupo_num = grupo_id[0]
                cuatrimestre = grupo_id[1]
                año = f"20{grupo_id[2:4]}"
                carrera_sigla = grupo_id[4:]
                
                # Determinar programa educativo basado en cuatrimestre
                # Programa 1: 7° y 10°
                # Programa 2: 1° y 4°
                if cuatrimestre in ['7', '10']:
                    programa_educativo = 1
                else:
                    programa_educativo = 2
                
                # Mapeo de siglas a nombres completos
                CARRERAS_PROGRAMA_1 = {
                    "IS": "Ingeniería en Software",
                    "IMA": "Ingeniería en Mecánica Automotriz",
                    "IF": "Ingeniería Financiera",
                    "ITM": "Ingeniería en Tecnologías de Manufactura",
                    "LNI": "Licenciatura en Negocios Internacionales"
                }
                
                CARRERAS_PROGRAMA_2 = {
                    "ITII": "Licenciatura en Ingeniería en Tecnologías de la Información e Innovación Digital",
                    "IMA": "Licenciatura en Ingeniería en Mecánica Automotriz",
                    "IF": "Licenciatura en Ingeniería Financiera",
                    "IMAV": "Licenciatura en Ingeniería en Manufactura Avanzadas",
                    "LCIA": "Licenciatura en Comercio Internacional y Aduanas"
                }
                
                if programa_educativo == 1:
                    carrera_nombre = CARRERAS_PROGRAMA_1.get(carrera_sigla, "No especificada")
                else:
                    carrera_nombre = CARRERAS_PROGRAMA_2.get(carrera_sigla, "No especificada")
                
                # Leer estudiantes de la hoja (empezando desde la fila 2)
                for row in range(2, sheet.max_row + 1):
                    nombre = sheet.cell(row, 1).value
                    apellido_p = sheet.cell(row, 2).value
                    apellido_m = sheet.cell(row, 3).value
                    matricula = sheet.cell(row, 4).value
                    
                    # Saltar filas vacías
                    if not nombre or not apellido_p or not matricula:
                        continue
                    
                    # Verificar si ya existe
                    cursor.execute("SELECT id FROM estudiantes WHERE matricula = ?", (str(matricula),))
                    existe = cursor.fetchone()
                    
                    if existe:
                        print(f"  ⚠️  Estudiante {nombre} {apellido_p} (Matrícula: {matricula}) ya existe. Omitiendo...")
                        estudiantes_duplicados += 1
                        continue
                    
                    # Insertar estudiante
                    cursor.execute(
                        '''
                        INSERT INTO estudiantes (matricula, nombre, apellido_p, apellido_m, cuatrimestre_actual, carrera, grupo, programa_educativo, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''',
                        (
                            str(matricula),
                            nombre,
                            apellido_p,
                            apellido_m or '',
                            cuatrimestre,
                            carrera_nombre,
                            grupo_id,
                            programa_educativo,
                            datetime.now().isoformat(),
                            datetime.now().isoformat()
                        )
                    )
                    
                    print(f"  ✓ Insertado: {nombre} {apellido_p} {apellido_m or ''} - Matrícula: {matricula} - Grupo: {grupo_id}")
                    estudiantes_insertados += 1
        
        conn.commit()
        print(f"\n✅ Carga completada:")
        print(f"   - Estudiantes insertados: {estudiantes_insertados}")
        print(f"   - Estudiantes duplicados (omitidos): {estudiantes_duplicados}")
        
    except Exception as e:
        print(f"❌ Error durante la carga: {e}")
        conn.rollback()
    finally:
        conn.close()

if __name__ == '__main__':
    load_data()
